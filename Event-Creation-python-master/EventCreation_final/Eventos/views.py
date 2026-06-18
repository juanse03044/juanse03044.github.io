from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, UpdateView, CreateView, DeleteView
from django.views import View
from django.utils.decorators import method_decorator
from functools import wraps
from django.db.models import Count
from django.http import HttpResponse
from django.contrib import messages
import json
import csv
import openpyxl

from .models import (
    Administradores,
    Eventos,
    Horarios,
    Invitados,
    Recursos,
    TipoEvento,
    Usuarios
)

from .forms import (
    LoginForm,
    RegistroForm,
    EditarPerfilForm,
    CambiarPasswordForm,
    TipoEventoForm,
    EventoForm,
    AdministradorForm,
    HorariosForm,
    InvitadosForm,
    RecursosForm,
    GestionUsuarioForm
)


# =========================
# 🔐 DECORADOR LOGIN
# =========================

def login_requerido(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


# =========================
# 🏠 INICIO
# =========================

def inicio(request):
    return render(request, 'index.html')


# =========================
# 🔐 LOGIN
# =========================

class LoginView(View):
    template_name = 'auth/login.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': LoginForm()
        })

    def post(self, request):
        form = LoginForm(request.POST)

        if form.is_valid():
            u = form.usuario
            request.session['usuario_id'] = u.pk
            request.session['usuario_rol'] = u.rol
            return redirect('inicio')

        return render(request, self.template_name, {
            'form': form
        })


def logout_view(request):
    request.session.flush()
    return redirect('login')


# =========================
# 📝 REGISTRO
# =========================

class RegistroView(View):
    template_name = 'auth/registro.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': RegistroForm()
        })

    def post(self, request):
        form = RegistroForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('login')

        return render(request, self.template_name, {
            'form': form
        })


# =========================
# 📋 LISTA USUARIOS
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaUsuarios(ListView):
    model = Usuarios
    template_name = 'usuarios/listaUsuarios.html'
    context_object_name = 'usuarios'


# =========================
# 📋 LISTA ADMINISTRADORES
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaAdministradores(ListView):
    model = Administradores
    template_name = 'administradores/listaAdministradores.html'
    context_object_name = 'administradores'


# =========================
# 📋 LISTA EVENTOS
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaEventos(ListView):
    model = Eventos
    template_name = 'eventos/listaEventos.html'
    context_object_name = 'eventos'

    def get_queryset(self):
        rol = self.request.session.get('usuario_rol')

        if rol == 'admin':
            # Admin ve todos los eventos
            return Eventos.objects.select_related('id_tipo').all()
        else:
            # Usuario ve solo los eventos donde está como invitado
            usuario_id = self.request.session.get('id_usuario')
            eventos_ids = Invitados.objects.filter(
                id_usuario=usuario_id
            ).values_list('evento_id', flat=True)
            return Eventos.objects.filter(
                id_evento__in=eventos_ids
            ).select_related('id_tipo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['es_admin'] = self.request.session.get('usuario_rol') == 'admin'
        return context


# =========================
# 📋 LISTA TIPO EVENTO
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaTipoEvento(ListView):
    model = TipoEvento
    template_name = 'eventos/listaTipoEvento.html'
    context_object_name = 'tipos'


# =========================
# 📋 LISTA HORARIOS
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaHorarios(ListView):
    model = Horarios
    template_name = 'horarios/listaHorarios.html'
    context_object_name = 'horarios'


# =========================
# 📋 LISTA INVITADOS
# =========================

@method_decorator(login_requerido, name='dispatch')
class ListaInvitados(ListView):
    model = Invitados
    template_name = 'invitados/listaInvitados.html'
    context_object_name = 'invitados'


# =========================
# 📋 LISTA RECURSOS
# =========================

@method_decorator(login_requerido, name='dispatch')
class listaRecursos(ListView):
    model = Recursos
    template_name = 'recursos/listaRecursos.html'
    context_object_name = 'recursos'


# =========================
# ➕ CREAR USUARIOS
# =========================

class createUsuarios(CreateView):
    model = Usuarios
    form_class = GestionUsuarioForm
    template_name = 'usuarios/crearUsuario.html'
    success_url = reverse_lazy('listaUsuarios')


# =========================
# ➕ CREAR ADMINISTRADORES
# =========================

class createAdministradores(CreateView):
    model = Administradores
    form_class = AdministradorForm
    template_name = 'administradores/crearAdministrador.html'
    success_url = reverse_lazy('listaAdministradores')


# =========================
# ➕ CREAR EVENTOS
# =========================

class createEventos(CreateView):
    model = Eventos
    form_class = EventoForm
    template_name = 'eventos/crearEvento.html'
    success_url = reverse_lazy('listaEventos')


# =========================
# ➕ CREAR TIPO EVENTO
# =========================

class createTipoEventos(CreateView):
    model = TipoEvento
    form_class = TipoEventoForm
    template_name = 'eventos/crearTipoEvento.html'
    success_url = reverse_lazy('listaTipoEvento')


# =========================
# ➕ CREAR HORARIOS
# =========================

class CreateHorarios(CreateView):
    model = Horarios
    form_class = HorariosForm
    template_name = 'horarios/crearHorario.html'
    success_url = reverse_lazy('listaHorarios')


# =========================
# ➕ CREAR INVITADOS
# =========================

class CreateInvitados(CreateView):
    model = Invitados
    form_class = InvitadosForm
    template_name = 'invitados/crearInvitado.html'
    success_url = reverse_lazy('listaInvitados')


# =========================
# ➕ CREAR RECURSOS
# =========================

class createRecursos(CreateView):
    model = Recursos
    form_class = RecursosForm
    template_name = 'recursos/crearRecurso.html'
    success_url = reverse_lazy('listaRecursos')


# =========================
# ✏️ EDITAR USUARIOS
# =========================

class updateUsuarios(UpdateView):
    model = Usuarios
    form_class = GestionUsuarioForm
    template_name = 'usuarios/crearUsuario.html'
    success_url = reverse_lazy('listaUsuarios')


# =========================
# ✏️ EDITAR ADMINISTRADORES
# =========================

class updateAdministradores(UpdateView):
    model = Administradores
    form_class = AdministradorForm
    template_name = 'administradores/crearAdministrador.html'
    success_url = reverse_lazy('listaAdministradores')


# =========================
# ✏️ EDITAR EVENTOS
# =========================

class updateEventos(UpdateView):
    model = Eventos
    form_class = EventoForm
    template_name = 'eventos/crearEvento.html'
    success_url = reverse_lazy('listaEventos')


# =========================
# ✏️ EDITAR TIPO EVENTO
# =========================

class updateTipoEventos(UpdateView):
    model = TipoEvento
    form_class = TipoEventoForm
    template_name = 'eventos/crearTipoEvento.html'
    success_url = reverse_lazy('listaTipoEvento')


# =========================
# ✏️ EDITAR HORARIOS
# =========================

class UpdateHorarios(UpdateView):
    model = Horarios
    form_class = HorariosForm
    template_name = 'horarios/crearHorario.html'
    success_url = reverse_lazy('listaHorarios')


# =========================
# ✏️ EDITAR INVITADOS
# =========================

class UpdateInvitados(UpdateView):
    model = Invitados
    form_class = InvitadosForm
    template_name = 'invitados/crearInvitado.html'
    success_url = reverse_lazy('listaInvitados')


# =========================
# ✏️ EDITAR RECURSOS
# =========================

class updateRecursos(UpdateView):
    model = Recursos
    form_class = RecursosForm
    template_name = 'recursos/crearRecurso.html'
    success_url = reverse_lazy('listaRecursos')


# =========================
# ❌ ELIMINAR USUARIOS
# =========================

class deleteUsuarios(DeleteView):
    model = Usuarios
    template_name = 'usuarios/eliminarUsuario.html'
    success_url = reverse_lazy('listaUsuarios')


# =========================
# ❌ ELIMINAR ADMINISTRADORES
# =========================

class deleteAdministradores(DeleteView):
    model = Administradores
    template_name = 'administradores/eliminarAdministrador.html'
    success_url = reverse_lazy('listaAdministradores')


# =========================
# ❌ ELIMINAR EVENTOS
# =========================

class deleteEventos(DeleteView):
    model = Eventos
    template_name = 'eventos/eliminarEvento.html'
    success_url = reverse_lazy('listaEventos')


# =========================
# ❌ ELIMINAR TIPO EVENTO
# =========================

class deleteTipoEvento(DeleteView):
    model = TipoEvento
    template_name = 'eventos/eliminarTipoEvento.html'
    success_url = reverse_lazy('listaTipoEvento')


# =========================
# ❌ ELIMINAR HORARIOS
# =========================

class deleteHorarios(DeleteView):
    model = Horarios
    template_name = 'horarios/eliminarHorario.html'
    success_url = reverse_lazy('listaHorarios')


# =========================
# ❌ ELIMINAR INVITADOS
# =========================

class EliminarInvitado(DeleteView):
    model = Invitados
    template_name = 'invitados/eliminarInvitado.html'
    success_url = reverse_lazy('listaInvitados')


# =========================
# ❌ ELIMINAR RECURSOS
# =========================

class deleteRecursos(DeleteView):
    model = Recursos
    template_name = 'recursos/eliminarRecurso.html'
    success_url = reverse_lazy('listaRecursos')


# =========================
# 📊 DASHBOARD
# =========================

@login_requerido
def dashboard(request):
    rol = request.session.get('usuario_rol')
    usuario_id = request.session.get('usuario_id')
    es_admin = rol == 'admin'

    if es_admin:
        total_eventos = Eventos.objects.count()
        total_invitados = Invitados.objects.count()
    else:
        # Solo cuenta los eventos del usuario
        eventos_ids = Invitados.objects.filter(
            usuario_id=usuario_id
        ).values_list('evento_id', flat=True)
        total_eventos = len(eventos_ids)
        total_invitados = Invitados.objects.filter(usuario_id=usuario_id).count()

    total_usuarios = Usuarios.objects.count() if es_admin else None
    total_recursos = Recursos.objects.count() if es_admin else None

    datos = TipoEvento.objects.annotate(total=Count('eventos'))
    labels = [d.nombre for d in datos]
    data = [d.total for d in datos]

    context = {
        'es_admin': es_admin,
        'total_eventos': total_eventos,
        'total_usuarios': total_usuarios,
        'total_invitados': total_invitados,
        'total_recursos': total_recursos,
        'labels': json.dumps(labels),
        'data': json.dumps(data),
    }
    return render(request, 'eventos/dashboard.html', context)


# =========================
# 👤 PERFIL USUARIO
# =========================

@login_requerido
def perfil_view(request):
    return render(request, 'usuarios/perfil.html')


# =========================
# ✏️ EDITAR PERFIL
# =========================

@login_requerido
def editar_perfil_view(request):
    usuario_id = request.session.get('usuario_id')
    usuario = Usuarios.objects.get(pk=usuario_id)

    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('perfil')
    else:
        form = EditarPerfilForm(instance=usuario)

    return render(request, 'usuarios/editar_perfil.html', {'form': form})


# =========================
# 🔑 CAMBIAR PASSWORD
# =========================

@login_requerido
def cambiar_password_view(request):
    usuario_id = request.session.get('usuario_id')
    usuario = Usuarios.objects.get(pk=usuario_id)

    if request.method == 'POST':
        form = CambiarPasswordForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('perfil')
    else:
        form = CambiarPasswordForm(instance=usuario)

    return render(request, 'usuarios/cambiar_password.html', {'form': form})


# =========================
# 📊 REPORTE EVENTOS
# =========================

@login_requerido
def reporte_eventos(request):
    eventos = Eventos.objects.select_related('id_tipo').all()
    return render(request, 'eventos/reporte_eventos.html', {'eventos': eventos})


# =========================
# 📥 EXPORTAR EXCEL
# =========================

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["ID", "Título", "Fecha", "Lugar", "Tipo"])

    for e in Eventos.objects.select_related('id_tipo').all():
        ws.append([
            e.id_evento,
            e.titulo,
            str(e.fecha),
            e.lugar,
            e.id_tipo.nombre
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="eventos.xlsx"'
    wb.save(response)
    return response


# =========================
# 📄 EXPORTAR PDF
# =========================

def exportar_pdf(request):
    from reportlab.pdfgen import canvas as rl_canvas

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="eventos.pdf"'

    p = rl_canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Lista de Eventos")

    y = 760
    p.setFont("Helvetica", 12)
    for e in Eventos.objects.select_related('id_tipo').all():
        linea = f"{e.id_evento} - {e.titulo} - {e.fecha} - {e.lugar} - {e.id_tipo.nombre}"
        p.drawString(100, y, linea)
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response


# =========================
# 📥 CARGA MASIVA
# =========================

def carga_masiva_general(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')

        if not archivo:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect('carga_masiva_general')

        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            creados = 0

            # Columnas esperadas en el Excel:
            # A: titulo | B: fecha (YYYY-MM-DD) | C: lugar | D: id_tipo (número)
            for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                titulo  = fila[0]
                fecha   = fila[1]
                lugar   = fila[2]
                id_tipo = fila[3]

                if not titulo:
                    errores.append(f'Fila {i}: título vacío')
                    continue

                try:
                    tipo = TipoEvento.objects.get(pk=id_tipo)
                except TipoEvento.DoesNotExist:
                    errores.append(f'Fila {i}: TipoEvento con id {id_tipo} no existe')
                    continue

                Eventos.objects.create(
                    titulo=titulo,
                    fecha=fecha,
                    lugar=lugar or '',
                    id_tipo=tipo
                )
                creados += 1

            if errores:
                messages.warning(
                    request,
                    f'Se crearon {creados} eventos. Errores: {", ".join(errores)}'
                )
            else:
                messages.success(request, f'Se cargaron {creados} eventos correctamente.')

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')

        return redirect('listaEventos')

    return render(request, 'eventos/carga_masiva_general.html')